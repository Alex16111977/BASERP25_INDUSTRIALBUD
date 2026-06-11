# -*- coding: utf-8 -*-
"""
Сверка актов списания стройматериалов (Глобино) между Бухгалтерией и ЕРП.

Источники (только чтение):
  1) Строительные материалы из Бухгалтерии.xlsx — листы "Глобино-2" (счёт 205)
     и "Глобино-2051" (счёт 2051), формат карточки счёта.
  2) Строительные материалы ЕРП.xlsx — лист "Лист2", секция "Строительные материалы",
     документы "Списание недостач товаров ВН-ІБ00-XXXXXX".

Ключ сверки: номер ІБ00-XXXXXX. ЕРП-сторона — только ВН-документы;
Бух-сторона — акты типов "Вимога-накладна" / "Списання товарів",
сумма = кредитовый оборот (списание со счёта), агрегация сквозь оба листа.

Результат: Сверка актов Глобино.xlsx (листы: Сверка, Бух прочие, Контроль).
Дизайн: docs/superpowers/specs/2026-06-12-sverka-aktov-globino-design.md
"""
import re
import sys
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры"
FILE_BUH = DIR + r"\Строительные материалы из Бухгалтерии.xlsx"
FILE_ERP = DIR + r"\Строительные материалы ЕРП.xlsx"
FILE_OUT = DIR + r"\Сверка актов Глобино.xlsx"

BUH_SHEETS = {"Глобино-2": "205", "Глобино-2051": "2051"}
ACT_TYPES = ("Вимога-накладна", "Списання товарів")

RX_NUM = re.compile(r"ІБ00-\d{6}")
RX_DATE = re.compile(r"(?:від|от)\s+(\d{2}\.\d{2}\.\d{4})")


def to_num(v):
    """Число из ячейки; строки из пробелов/пустые -> 0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ----------------------------------------------------------------------------
# 1. ЕРП: секция "Строительные материалы"
# ----------------------------------------------------------------------------
def read_erp():
    wb = openpyxl.load_workbook(FILE_ERP, read_only=True, data_only=True)
    ws = wb["Лист2"]
    acts = {}          # номер -> dict(дата, текст, сумма)
    realiz_total = 0.0
    grand_total = None
    section = None     # None -> "realiz" -> "stroy"
    for row in ws.iter_rows(values_only=True):
        text = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        amount = to_num(row[5]) if len(row) > 5 else 0.0
        if text == "Себестоимость реализации":
            section = "realiz"
            continue
        if text == "Строительные материалы":
            section = "stroy"
            continue
        if not text:
            if amount:
                grand_total = amount  # итоговая строка листа (без текста)
            continue
        m = RX_NUM.search(text)
        if section == "realiz" and "Реализация" in text:
            realiz_total += amount
            continue
        if section == "stroy" and m:
            if "ВН-" not in text:
                print(f"  ! ЕРП: документ без префикса ВН- пропущен: {text[:80]}")
                continue
            num = m.group(0)
            dm = RX_DATE.search(text)
            if num in acts:
                raise RuntimeError(f"ЕРП: дубль номера {num}")
            acts[num] = {
                "date": dm.group(1) if dm else "",
                "text": text,
                "sum": amount,
            }
    wb.close()
    return acts, realiz_total, grand_total


# ----------------------------------------------------------------------------
# 2. Бухгалтерия: листы Глобино
# ----------------------------------------------------------------------------
def parse_doc(cell):
    """'Вимога-накладна ІБ00-000285 від ... 18:00:11\\nСписаны ТМЦ' ->
    (тип, номер, дата, операция)."""
    s = str(cell)
    lines = s.split("\n")
    head = lines[0].strip()
    oper = lines[1].strip() if len(lines) > 1 else ""
    m = RX_NUM.search(head)
    num = m.group(0) if m else ""
    dtype = head[: m.start()].strip() if m else head
    dm = RX_DATE.search(head)
    return dtype, num, (dm.group(1) if dm else ""), oper


def read_buh():
    wb = openpyxl.load_workbook(FILE_BUH, read_only=True, data_only=True)
    # акты: номер -> {"type", "date", "205": кт-сумма, "2051": кт-сумма, "dt": дт-сумма}
    acts = defaultdict(lambda: {"type": "", "date": "", "205": 0.0, "2051": 0.0, "dt": 0.0})
    # прочие: (лист, тип, номер, операция) -> {"date", "dt", "kt", "rows"}
    others = defaultdict(lambda: {"date": "", "dt": 0.0, "kt": 0.0, "rows": 0})
    sheet_totals = {}   # лист -> {"dt", "kt", "rows"}
    collisions = []

    for sheet, account in BUH_SHEETS.items():
        ws = wb[sheet]
        tot_dt = tot_kt = 0.0
        rows = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            doc = row[1] if len(row) > 1 else None
            if not doc:
                continue
            dt_sum = to_num(row[5]) if len(row) > 5 else 0.0
            kt_sum = to_num(row[8]) if len(row) > 8 else 0.0
            dtype, num, date, oper = parse_doc(doc)
            rows += 1
            tot_dt += dt_sum
            tot_kt += kt_sum
            if dtype in ACT_TYPES and num:
                rec = acts[num]
                if rec["type"] and rec["type"] != dtype:
                    collisions.append((num, rec["type"], dtype))
                rec["type"] = dtype
                rec["date"] = rec["date"] or date
                rec[account] += kt_sum
                rec["dt"] += dt_sum
            else:
                key = (sheet, dtype, num, oper)
                o = others[key]
                o["date"] = o["date"] or date
                o["dt"] += dt_sum
                o["kt"] += kt_sum
                o["rows"] += 1
        sheet_totals[sheet] = {"dt": tot_dt, "kt": tot_kt, "rows": rows}
    wb.close()
    return acts, others, sheet_totals, collisions


# ----------------------------------------------------------------------------
# 3. Сверка
# ----------------------------------------------------------------------------
def build_compare(erp_acts, buh_acts):
    nums = sorted(set(erp_acts) | set(buh_acts))
    out = []
    for num in nums:
        e = erp_acts.get(num)
        b = buh_acts.get(num)
        s_erp = e["sum"] if e else None
        s205 = b["205"] if b else None
        s2051 = b["2051"] if b else None
        s_buh = (s205 or 0.0) + (s2051 or 0.0) if b else None
        ratio = None
        if e and b:
            diff = s_erp - s_buh
            ratio = (s_erp / s_buh) if s_buh else None
            if abs(diff) < 0.01:
                status = "OK"
            elif abs(s_erp - s_buh * 1.2) <= 0.5:
                # ЕРП хранит себестоимость с НДС 20%, Бух — без НДС;
                # допуск 0.5 грн на построчные округления
                status = "OK (НДС 20%)"
            else:
                status = "Расхождение"
        elif e:
            diff, status = None, "Только в ЕРП"
        else:
            diff, status = None, "Только в Бух"
        out.append({
            "num": num,
            "date": (e["date"] if e else b["date"]),
            "erp_doc": ("Списание недостач товаров ВН-" + num if e else ""),
            "erp_sum": s_erp,
            "buh_doc": (b["type"] + " " + num if b else ""),
            "buh205": s205,
            "buh2051": s2051,
            "buh_sum": s_buh,
            "diff": diff,
            "ratio": ratio,
            "status": status,
        })
    return out


# ----------------------------------------------------------------------------
# 4. Выгрузка в Excel
# ----------------------------------------------------------------------------
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
F_HDR = Font(bold=True, size=10)
F_TOT = Font(bold=True)
FILL_HDR = PatternFill("solid", fgColor="DDEBF7")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
NUMFMT = "# ##0.00"


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def write_output(compare, erp_acts, buh_acts, others, sheet_totals,
                 realiz_total, grand_total, collisions):
    wb = openpyxl.Workbook()

    # --- Лист 1: Сверка -----------------------------------------------------
    ws = wb.active
    ws.title = "Сверка"
    headers = ["№", "Номер", "Дата", "Документ ЕРП", "Сумма ЕРП",
               "Документ Бух", "Бух 205", "Бух 2051", "Бух итого",
               "Разница ЕРП-Бух", "ЕРП/Бух", "Статус"]
    ws.append(headers)
    style_header(ws, len(headers))
    for i, r in enumerate(compare, start=1):
        ws.append([i, r["num"], r["date"], r["erp_doc"], r["erp_sum"],
                   r["buh_doc"], r["buh205"], r["buh2051"], r["buh_sum"],
                   r["diff"], r["ratio"], r["status"]])
        row = ws.max_row
        for c in (5, 7, 8, 9, 10):
            ws.cell(row=row, column=c).number_format = NUMFMT
        ws.cell(row=row, column=11).number_format = "0.0000"
        st = r["status"]
        fill = (FILL_OK if st in ("OK", "OK (НДС 20%)")
                else FILL_BAD if st == "Расхождение" else FILL_WARN)
        ws.cell(row=row, column=12).fill = fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN
    # итоги
    tot_erp = sum(r["erp_sum"] or 0 for r in compare)
    tot205 = sum(r["buh205"] or 0 for r in compare)
    tot2051 = sum(r["buh2051"] or 0 for r in compare)
    tot_buh = sum(r["buh_sum"] or 0 for r in compare)
    ws.append(["", "ИТОГО", "", "", tot_erp, "", tot205, tot2051, tot_buh,
               tot_erp - tot_buh, "", ""])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = F_TOT
        if c in (5, 7, 8, 9, 10):
            cell.number_format = NUMFMT
    widths = [5, 13, 11, 42, 14, 30, 13, 13, 14, 14, 9, 15]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(compare) + 1}"

    # --- Лист 2: Бух прочие документы ---------------------------------------
    ws2 = wb.create_sheet("Бух прочие")
    h2 = ["Лист", "Тип документа", "Номер", "Дата", "Операция",
          "Строк", "Дебет (приход)", "Кредит (расход)"]
    ws2.append(h2)
    style_header(ws2, len(h2))
    for (sheet, dtype, num, oper) in sorted(others):
        o = others[(sheet, dtype, num, oper)]
        ws2.append([sheet, dtype, num, o["date"], oper, o["rows"], o["dt"], o["kt"]])
        for c in (7, 8):
            ws2.cell(row=ws2.max_row, column=c).number_format = NUMFMT
    ws2.append(["ИТОГО", "", "", "", "",
                sum(o["rows"] for o in others.values()),
                sum(o["dt"] for o in others.values()),
                sum(o["kt"] for o in others.values())])
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(row=ws2.max_row, column=c)
        cell.font = F_TOT
        if c in (7, 8):
            cell.number_format = NUMFMT
    for c, w in enumerate([14, 28, 13, 11, 34, 7, 15, 15], start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{1 + len(others)}"

    # --- Лист 3: Контроль ----------------------------------------------------
    ws3 = wb.create_sheet("Контроль")
    ws3.append(["Показатель", "Значение"])
    style_header(ws3, 2)

    korr = {k: v for k, v in others.items() if "Корректировка стоимости списания" in k[3]}
    korr_kt = sum(v["kt"] for v in korr.values())
    korr_dt = sum(v["dt"] for v in korr.values())

    rows = [
        ("ЕРП: документов в секции 'Строительные материалы'", len(erp_acts)),
        ("ЕРП: сумма секции 'Строительные материалы'", tot_erp),
        ("ЕРП: сумма секции 'Себестоимость реализации' (вне сверки)", realiz_total),
        ("ЕРП: общий итог листа (последняя строка файла)", grand_total),
        ("КОНТРОЛЬ ЕРП: секция + реализации - итог листа (должно быть ~0)",
         tot_erp + realiz_total - (grand_total or 0)),
        ("", ""),
        ("Бух: актов (Вимога-накладна/Списання товарів)", len(buh_acts)),
        ("Бух: сумма актов Кт, счёт 205", tot205),
        ("Бух: сумма актов Кт, счёт 2051", tot2051),
        ("Бух: сумма актов Кт, всего", tot_buh),
        ("", ""),
    ]
    for sheet, t in sheet_totals.items():
        rows.append((f"Бух лист '{sheet}': строк проводок", t["rows"]))
        rows.append((f"Бух лист '{sheet}': оборот Дт", t["dt"]))
        rows.append((f"Бух лист '{sheet}': оборот Кт", t["kt"]))
    kt_all = sum(t["kt"] for t in sheet_totals.values())
    kt_others = sum(o["kt"] for o in others.values())
    rows += [
        ("", ""),
        ("КОНТРОЛЬ Бух: Кт актов + Кт прочих - Кт листов (должно быть 0)",
         tot_buh + kt_others - kt_all),
        ("", ""),
        ("Закрытие месяца: корректировка стоимости списания, Кт", korr_kt),
        ("Закрытие месяца: корректировка стоимости списания, Дт", korr_dt),
        ("", ""),
        ("Сверка: всего номеров", len(compare)),
        ("Сверка: OK (суммы равны)", sum(1 for r in compare if r["status"] == "OK")),
        ("Сверка: OK (НДС 20%: ЕРП = Бух x 1.2)",
         sum(1 for r in compare if r["status"] == "OK (НДС 20%)")),
        ("Сверка: Расхождение", sum(1 for r in compare if r["status"] == "Расхождение")),
        ("Сверка: Только в ЕРП", sum(1 for r in compare if r["status"] == "Только в ЕРП")),
        ("Сверка: Только в Бух", sum(1 for r in compare if r["status"] == "Только в Бух")),
        ("Сверка: разница итогов ЕРП - Бух", tot_erp - tot_buh),
        ("", ""),
        ("Справочно: Бух по совпавшим актам x 1.2 (как было бы с НДС)",
         sum((r["buh_sum"] or 0) * 1.2 for r in compare if r["erp_sum"] is not None and r["buh_sum"] is not None)),
        ("Справочно: ЕРП по совпавшим актам",
         sum(r["erp_sum"] or 0 for r in compare if r["buh_sum"] is not None)),
    ]
    if collisions:
        rows.append(("ВНИМАНИЕ: коллизии типов актов по номеру", str(collisions)))
    for name, val in rows:
        ws3.append([name, val])
        if isinstance(val, float):
            ws3.cell(row=ws3.max_row, column=2).number_format = NUMFMT
    ws3.column_dimensions["A"].width = 62
    ws3.column_dimensions["B"].width = 20

    wb.save(FILE_OUT)
    return tot_erp, tot_buh, tot205, tot2051


# ----------------------------------------------------------------------------
def main():
    print("1. Читаю ЕРП ...")
    erp_acts, realiz_total, grand_total = read_erp()
    print(f"   актов ЕРП: {len(erp_acts)}, сумма: {sum(a['sum'] for a in erp_acts.values()):,.2f}")
    print(f"   реализации (вне сверки): {realiz_total:,.2f}; итог листа: {grand_total:,.2f}")

    print("2. Читаю Бухгалтерию ...")
    buh_acts, others, sheet_totals, collisions = read_buh()
    print(f"   актов Бух: {len(buh_acts)}, прочих документов: {len(others)}")
    if collisions:
        print(f"   ! КОЛЛИЗИИ типов по номеру: {collisions}")
    dt_in_acts = sum(a["dt"] for a in buh_acts.values())
    if abs(dt_in_acts) > 0.005:
        print(f"   ! у актов есть дебетовые суммы (ожидалось 0): {dt_in_acts:,.2f}")

    print("3. Сверяю ...")
    compare = build_compare(erp_acts, buh_acts)

    print("4. Пишу результат ...")
    tot_erp, tot_buh, tot205, tot2051 = write_output(
        compare, erp_acts, buh_acts, others, sheet_totals,
        realiz_total, grand_total, collisions)

    # Контрольные суммы в консоль
    ctrl_erp = tot_erp + realiz_total - (grand_total or 0)
    kt_all = sum(t["kt"] for t in sheet_totals.values())
    kt_others = sum(o["kt"] for o in others.values())
    ctrl_buh = tot_buh + kt_others - kt_all
    print("-" * 70)
    print(f"ИТОГ ЕРП (стройматериалы):  {tot_erp:>16,.2f}")
    print(f"ИТОГ Бух (акты 205+2051):   {tot_buh:>16,.2f}  (205: {tot205:,.2f} / 2051: {tot2051:,.2f})")
    print(f"Разница ЕРП-Бух:            {tot_erp - tot_buh:>16,.2f}")
    print(f"КОНТРОЛЬ ЕРП (~0):          {ctrl_erp:>16,.2f}")
    print(f"КОНТРОЛЬ Бух (=0):          {ctrl_buh:>16,.2f}")
    for st in ("OK", "OK (НДС 20%)", "Расхождение", "Только в ЕРП", "Только в Бух"):
        n = sum(1 for r in compare if r["status"] == st)
        print(f"  {st:<14}: {n}")
    if abs(ctrl_erp) > 0.02 or abs(ctrl_buh) > 0.005:
        print("!!! КОНТРОЛЬНЫЕ СУММЫ НЕ СОШЛИСЬ — результат не использовать")
        sys.exit(1)
    print(f"OK -> {FILE_OUT}")


if __name__ == "__main__":
    main()
