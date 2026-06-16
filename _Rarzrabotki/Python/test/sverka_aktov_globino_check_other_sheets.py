# -*- coding: utf-8 -*-
"""Ищем номера 000329, 000354, 000346, 000347 по ВСЕМ листам Бух-файла."""
import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, re

BUH = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры\Строительные материалы из Бухгалтерии.xlsx"
TARGETS = ["ІБ00-000329", "ІБ00-000354", "ІБ00-000346", "ІБ00-000347"]

wb = openpyxl.load_workbook(BUH, read_only=True, data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        doc = row[1]
        if not doc:
            continue
        for t in TARGETS:
            if t in str(doc):
                head = str(doc).split("\n")[0]
                kt = row[8] if len(row) > 8 else None
                print(f"{sheet} r{i}: {head} | Кт={kt}")
wb.close()
print("DONE")
