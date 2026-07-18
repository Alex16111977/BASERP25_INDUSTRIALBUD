# -*- coding: utf-8 -*-
"""Smoke внешней обработки «Загрузка материалов в СС» на базе BuhBud.
Ожидание вычисляется из того же Excel через openpyxl, факт — из ТЧ Комплектующие
временной СтруктураСебестоимости после вызова движка ЗагрузитьМатериалыИзМассива."""
import sys, os
sys.stdout.reconfigure(encoding="utf-8")
import win32com.client
import openpyxl

EPF = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Загрузка СС\Загрузка материалов в СС.epf"
XLSX = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Загрузка СС\Виробництво\1\IRC 15м2 НОВИЙ ШАблон  СС Виробництво 16-06-2026_Коррект (1).xlsx"
TESTNAME = "__ТЕСТ Загрузка материалов СС"
CONN = 'Srvr="localhost";Ref="bas_industrialbud";Usr="cfo";Pwd="2442"'

def num(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

def read_expected(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Проект СС"]
    def g(r, c): return ws.cell(row=r, column=c).value
    rows = []
    for r in range(2, ws.max_row + 1):
        c = g(r, 3)
        if (str(c).strip() if c is not None else "") != "Матеріал":
            continue
        gname = (str(g(r, 7)).strip() if g(r, 7) is not None else "")
        if gname == "":
            continue
        rows.append({
            "Этап": (str(g(r, 5)).strip() if g(r, 5) is not None else ""),
            "ОбщееНазвание": (str(g(r, 8)).strip() if g(r, 8) is not None else ""),
            "НоменклатураСС": gname,
            "ЕдиницаСС": (str(g(r, 9)).strip() if g(r, 9) is not None else ""),
            "Количество": num(g(r, 10)),
            "Сумма": num(g(r, 15)),
        })
    return rows

def get_or_create_struct(erp):
    q = erp.NewObject("Запрос")
    q.Text = "ВЫБРАТЬ Ссылка ИЗ Справочник.СтруктураСебестоимости ГДЕ Наименование = &Н"
    q.УстановитьПараметр("Н", TESTNAME)
    sel = q.Execute().Выбрать()
    if sel.Следующий():
        return sel.Ссылка
    obj = erp.Справочники.СтруктураСебестоимости.СоздатьЭлемент()
    obj.Наименование = TESTNAME
    obj.Записать()
    return obj.Ссылка

def build_array(erp, rows):
    mass = erp.NewObject("Массив")
    for row in rows:
        s = erp.NewObject("Структура")
        s.Вставить("Этап", row["Этап"])
        s.Вставить("ОбщееНазвание", row["ОбщееНазвание"])
        s.Вставить("НоменклатураСС", row["НоменклатураСС"])
        s.Вставить("ЕдиницаСС", row["ЕдиницаСС"])
        s.Вставить("Количество", row["Количество"])
        s.Вставить("Сумма", row["Сумма"])
        mass.Добавить(s)
    return mass

def tc_totals(erp, ref):
    q = erp.NewObject("Запрос")
    q.Text = ("ВЫБРАТЬ КОЛИЧЕСТВО(*) КАК К, СУММА(Сумма) КАК С, "
              "СУММА(ВЫБОР КОГДА Этап <> ЗНАЧЕНИЕ(Справочник.ЭтапыРабот.ПустаяСсылка) ТОГДА 1 ИНАЧЕ 0 КОНЕЦ) КАК СЭтап, "
              "СУММА(ВЫБОР КОГДА ОбщееНазвание <> ЗНАЧЕНИЕ(Справочник.ОбщиеНазванияНоменклатуры.ПустаяСсылка) ТОГДА 1 ИНАЧЕ 0 КОНЕЦ) КАК СНазв, "
              "СУММА(ВЫБОР КОГДА НоменклатураСС <> \"\" ТОГДА 1 ИНАЧЕ 0 КОНЕЦ) КАК СНаим "
              "ИЗ Справочник.СтруктураСебестоимости.Комплектующие ГДЕ Ссылка = &Р")
    q.УстановитьПараметр("Р", ref)
    sel = q.Execute().Выбрать()
    sel.Следующий()
    return int(sel.К or 0), float(sel.С or 0), int(sel.СЭтап or 0), int(sel.СНазв or 0), int(sel.СНаим or 0)

def tc_mismatch(erp, ref):
    """Инвариант: Номенклатура/Единица строк ТЧ == реквизиты ОбщегоНазвания. Возвращает (расхождений, строк с заполненной Номенклатурой)."""
    q = erp.NewObject("Запрос")
    q.Text = ("ВЫБРАТЬ КОЛИЧЕСТВО(*) КАК Расхожд "
              "ИЗ Справочник.СтруктураСебестоимости.Комплектующие КАК К "
              "ГДЕ К.Ссылка = &Р И (К.Номенклатура <> К.ОбщееНазвание.ОсновнаяНоменклатура "
              "ИЛИ К.Единица <> К.ОбщееНазвание.Единица)")
    q.УстановитьПараметр("Р", ref)
    sel = q.Execute().Выбрать()
    sel.Следующий()
    mism = int(sel.Расхожд or 0)
    q2 = erp.NewObject("Запрос")
    q2.Text = ("ВЫБРАТЬ КОЛИЧЕСТВО(*) КАК СНом "
               "ИЗ Справочник.СтруктураСебестоимости.Комплектующие "
               "ГДЕ Ссылка = &Р И Номенклатура <> ЗНАЧЕНИЕ(Справочник.Номенклатура.ПустаяСсылка)")
    q2.УстановитьПараметр("Р", ref)
    sel2 = q2.Execute().Выбрать()
    sel2.Следующий()
    return mism, int(sel2.СНом or 0)

def main():
    assert os.path.exists(EPF), f"НЕ найден .epf: {EPF} (собери обработку — Task 5)"
    exp = read_expected(XLSX)
    exp_cnt = len(exp)
    exp_sum = round(sum(r["Сумма"] for r in exp), 2)
    exp_nazv = sum(1 for r_ in exp if r_["ОбщееНазвание"] != "")
    print(f"ОЖИДАНИЕ (Excel): строк={exp_cnt}, ΣСумма={exp_sum}")

    v8 = win32com.client.Dispatch("V83.COMConnector")
    erp = v8.Connect(CONN)
    ref = get_or_create_struct(erp)
    mass = build_array(erp, exp)

    obr = erp.ВнешниеОбработки.Создать(EPF, False)
    res = obr.ЗагрузитьМатериалыИзМассива(ref, mass)
    print(f"ПРОТОКОЛ: Ошибка={res.Ошибка} Текст='{res.ТекстОшибки}' "
          f"Загружено={res.Загружено} СозданоЭтапов={res.СозданоЭтапов} "
          f"СозданоНазваний={res.СозданоНазваний} Пропущено={res.Пропущено}")
    assert not res.Ошибка, f"движок вернул ошибку: {res.ТекстОшибки}"
    assert int(res.Загружено) == exp_cnt, f"Загружено {res.Загружено} != {exp_cnt}"
    assert int(res.Пропущено) == 0, f"Пропущено {res.Пропущено} != 0 (массив предфильтрован)"

    k, s, s_etap, s_nazv, s_naim = tc_totals(erp, ref)
    print(f"ФАКТ (ТЧ): строк={k}, ΣСумма={round(s,2)}, сЭтапом={s_etap}, сНазв={s_nazv}, сНаим={s_naim}")
    assert k == exp_cnt, f"строк в ТЧ {k} != {exp_cnt}"
    assert abs(s - exp_sum) < 0.05, f"ΣСумма {s} != {exp_sum}"
    assert s_naim == exp_cnt, "не все НоменклатураСС заполнены"
    assert s_etap == exp_cnt, "не все Этап заполнены (ссылка)"
    assert s_nazv == exp_nazv, f"ОбщееНазвание заполнено {s_nazv} != {exp_nazv}"

    mism, s_nom = tc_mismatch(erp, ref)
    print(f"СВЯЗКА: строк с Номенклатурой={s_nom}, расхождений с ОбщимНазванием={mism}")
    assert mism == 0, f"Номенклатура/Единица не совпадают с реквизитами ОбщегоНазвания: {mism} строк"

    assert res.КонтрольПройден, "контроль загрузки не пройден"
    assert res.Контроль.Количество() > 0, "контрольный блок пуст"
    print("КОНТРОЛЬ:")
    for i in range(res.Контроль.Количество()):
        print("  " + str(res.Контроль.Получить(i)))
    for i in range(res.Инфо.Количество()):
        print("  " + str(res.Инфо.Получить(i)))

    # идемпотентность: повторный прогон — те же строки, 0 новых этапов
    res2 = obr.ЗагрузитьМатериалыИзМассива(ref, mass)
    assert not res2.Ошибка, f"повторный прогон вернул ошибку: {res2.ТекстОшибки}"
    assert res2.КонтрольПройден, "контроль на повторе не пройден"
    assert int(res2.Загружено) == exp_cnt, f"идемпотентность: загружено {res2.Загружено} != {exp_cnt}"
    k2, s2, _, _, _ = tc_totals(erp, ref)
    assert k2 == exp_cnt, f"идемпотентность: строк {k2} != {exp_cnt}"
    assert abs(s2 - exp_sum) < 0.05, f"идемпотентность: ΣСумма {s2} != {exp_sum}"
    assert int(res2.СозданоЭтапов) == 0, f"идемпотентность: создано этапов {res2.СозданоЭтапов} != 0"
    assert int(res2.СозданоНазваний) == 0, f"идемпотентность: создано названий {res2.СозданоНазваний} != 0"
    print("OK: smoke пройден (загрузка + суммы + этапы + идемпотентность)")

if __name__ == "__main__":
    main()
