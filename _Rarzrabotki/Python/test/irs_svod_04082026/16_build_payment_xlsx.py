# -*- coding: utf-8 -*-
"""Платіжний календар серпня по МД IRS 2026: скільки реально платити за матеріали.

Правильна сума серпня по ЗН = неоплачена кредиторка (вже отримано, не сплачено)
                            + залишок вартості того, що ще не куплено.
"""
import json, sys, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

OUT = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\А_ПланФактныйПроизводствоПолный\Бюджет"
raw = json.load(open("data_pay_raw.json", encoding="utf-8"))
byON = {d["ОН"]: d for d in json.load(open("data_byon_calc.json", encoding="utf-8"))}
ar = {x["ЗН"]: x for x in json.load(open("data_august_risk.json", encoding="utf-8"))["rows"]}

lines, debt, pays = raw["lines"], raw["debt"], raw["pays"]

# ---- агрегація по ЗН ----
zn = {}
for ln in lines:
    share = (ln["СумаЗН"] / ln["СумаДок"]) if ln["СумаДок"] else 0.0
    d_doc = debt.get(ln["ДокКлюч"], 0.0)
    unpaid = round(d_doc * share, 2)
    ln["Частка"] = round(share, 6)
    ln["БоргДок"] = d_doc
    ln["НеоплаченоЗН"] = unpaid
    ln["ОплаченоЗН"] = round(ln["СумаЗН"] - unpaid, 2)
    z = zn.setdefault(ln["ЗН"], {"Прихід": 0.0, "Оплачено": 0.0, "Неоплачено": 0.0, "накл": []})
    z["Прихід"] += ln["СумаЗН"]; z["Оплачено"] += ln["ОплаченоЗН"]; z["Неоплачено"] += unpaid
    z["накл"].append(ln)

rows = []
all_zn = set(zn) | set(byON)
for name in all_zn:
    z = zn.get(name, {"Прихід": 0.0, "Оплачено": 0.0, "Неоплачено": 0.0, "накл": []})
    b = byON.get(name, {})
    ost = b.get("Осталось", 0.0)               # ще не куплено (ETC)
    graf = ar.get(name, {}).get("ГрафікСерпень", 0.0)
    unpaid = round(z["Неоплачено"], 2)
    correct = round(unpaid + ost, 2)           # правильна потреба в грошах
    to_pay = round(min(graf, correct), 2) if graf else 0.0
    over = round(max(graf - correct, 0.0), 2)
    short = round(max(correct - graf, 0.0), 2)
    if graf <= 0.004:
        st = "немає у графіку серпня"
    elif over > 0.005 and correct <= 0.005:
        st = "ЗНЯТИ повністю"
    elif over > 0.005:
        st = "зменшити"
    elif short > 0.005:
        st = "недобір у графіку"
    else:
        st = "залишити як є"
    rows.append({"ЗН": name, "Од": b.get("Единица", ""), "Прихід": round(z["Прихід"], 2),
                 "Оплачено": round(z["Оплачено"], 2), "Неоплачено": unpaid,
                 "ЩеНеКуплено": ost, "ПравильнаСума": correct, "ГрафікСерпня": graf,
                 "ДоСплати": to_pay, "Зайве": over, "Недобір": short, "Статус": st,
                 "накл": z["накл"]})
rows.sort(key=lambda r: (-r["Зайве"], -r["ГрафікСерпня"]))

T = {k: round(sum(r[k] for r in rows), 2) for k in
     ("Прихід", "Оплачено", "Неоплачено", "ЩеНеКуплено", "ПравильнаСума",
      "ГрафікСерпня", "ДоСплати", "Зайве", "Недобір")}
print("ГрафікСерпня=%(ГрафікСерпня, ),.2f" % {} if False else
      "Графік серпня = {ГрафікСерпня:,.2f}\nПравильна сума = {ПравильнаСума:,.2f}\n"
      "  у т.ч. неоплачена кредиторка = {Неоплачено:,.2f}\n"
      "  у т.ч. ще не куплено         = {ЩеНеКуплено:,.2f}\n"
      "Зайве в графіку = {Зайве:,.2f}\nНедобір = {Недобір:,.2f}".format(**T))

# ================= Excel =================
HDR = PatternFill("solid", fgColor="1F4E79"); HF = Font(color="FFFFFF", bold=True, size=10)
MON = PatternFill("solid", fgColor="8A5A00")
L1P = PatternFill("solid", fgColor="DDEBF7"); L1F = Font(bold=True)
PAYF = Font(italic=True, size=9, color="1B7F4B")
INVF = Font(size=10)
RED = Font(color="B3261E", bold=True); AMB = Font(color="8A5A00")
TOTP = PatternFill("solid", fgColor="EEF3F8")
THIN = Border(bottom=Side(style="thin", color="C9D2DA"))
N2 = "#,##0.00"; N3 = "#,##0.###"

wb = openpyxl.Workbook(); wb.remove(wb.active)

def head(ws, headers, widths, mon_cols=()):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = MON if c in mon_cols else HDR
        cell.font = HF
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 42
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

# ---- 1. Зведення ----
ws = wb.create_sheet("Зведення")
data = [
    ("МД IRS 2026 — скільки платити за матеріали в серпні 2026", None),
    ("", None),
    ("ЛОГІКА", None),
    ("Колонка «Серпень» у кошторисі = гроші до сплати, а не план закупівлі.", None),
    ("Правильна сума = неоплачена кредиторка + вартість ще не купленого.", None),
    ("", None),
    ("ЩО ВЖЕ КУПЛЕНО (прихід ТМЦ, з ПДВ)", None),
    ("Прихід матеріалів разом", T["Прихід"]),
    ("з них оплачено", T["Оплачено"]),
    ("з них НЕ оплачено (кредиторка)", T["Неоплачено"]),
    ("", None),
    ("ЩО ЩЕ ТРЕБА КУПИТИ", None),
    ("Залишок до закупівлі за планом СС", T["ЩеНеКуплено"]),
    ("", None),
    ("ПОТРЕБА В ГРОШАХ", None),
    ("Правильна сума до сплати (кредиторка + закупівля)", T["ПравильнаСума"]),
    ("Стоїть у графіку серпня (МАТЕРИАЛИ.xlsx)", T["ГрафікСерпня"]),
    ("ЗНЯТИ з серпня (зайве)", T["Зайве"]),
    ("ДОДАТИ (недобір у графіку)", T["Недобір"]),
    ("", None),
    ("Реально до сплати в межах графіка", T["ДоСплати"]),
]
for a, b in data:
    ws.append([a, b])
ws.column_dimensions["A"].width = 58; ws.column_dimensions["B"].width = 20
for r in range(1, ws.max_row + 1):
    ws.cell(r, 2).number_format = N2
    v = ws.cell(r, 1).value or ""
    if v and ws.cell(r, 2).value is None and v == v.upper() and len(v) > 3:
        ws.cell(r, 1).font = Font(bold=True, color="1F4E79")
ws.cell(1, 1).font = Font(bold=True, size=13, color="1F4E79")
for r in (18, 19):
    ws.cell(r, 1).font = RED; ws.cell(r, 2).font = RED

# ---- 2. По ЗН ----
ws2 = wb.create_sheet("Скільки платити по ЗН")
head(ws2, ["Загальна назва", "Од.", "Прихід, грн", "Оплачено, грн", "НЕоплачено (кредиторка), грн",
           "Ще не куплено, грн", "ПРАВИЛЬНА сума серпня, грн", "Стоїть у графіку, грн",
           "Зайве (зняти), грн", "Недобір (додати), грн", "Статус"],
     [34, 8, 14, 14, 17, 15, 18, 16, 15, 16, 22], mon_cols=(7, 8, 9, 10))
for r in rows:
    ws2.append([r["ЗН"], r["Од"], r["Прихід"], r["Оплачено"], r["Неоплачено"], r["ЩеНеКуплено"],
                r["ПравильнаСума"], r["ГрафікСерпня"], r["Зайве"], r["Недобір"], r["Статус"]])
    rr = ws2.max_row
    for c in range(1, 12):
        ws2.cell(rr, c).border = THIN
        if c >= 3 and c <= 10: ws2.cell(rr, c).number_format = N2
    if r["Статус"].startswith("ЗНЯТИ"):
        for c in range(1, 12): ws2.cell(rr, c).font = RED
    elif r["Статус"].startswith("зменшити"):
        for c in range(1, 12): ws2.cell(rr, c).font = AMB
ws2.append(["РАЗОМ", "", T["Прихід"], T["Оплачено"], T["Неоплачено"], T["ЩеНеКуплено"],
            T["ПравильнаСума"], T["ГрафікСерпня"], T["Зайве"], T["Недобір"], ""])
for c in range(1, 12):
    ws2.cell(ws2.max_row, c).font = Font(bold=True); ws2.cell(ws2.max_row, c).fill = TOTP
    if 3 <= c <= 10: ws2.cell(ws2.max_row, c).number_format = N2
ws2.freeze_panes = "A2"; ws2.auto_filter.ref = "A1:K%d" % ws2.max_row

# ---- 3. Розшифровка: прихід + оплати ----
ws3 = wb.create_sheet("Розшифровка приходів і оплат")
head(ws3, ["Загальна назва / Накладна / Оплата", "Дата", "Постачальник", "Кіл-сть",
           "Сума позиції, грн", "Сума накладної, грн", "Оплачено, грн", "НЕоплачено, грн", "Тип"],
     [58, 12, 30, 10, 16, 18, 15, 16, 22])
for r in rows:
    if not r["накл"]:
        continue
    ws3.append([r["ЗН"], "", "", "", r["Прихід"], "", r["Оплачено"], r["Неоплачено"],
                "РАЗОМ ПО ЗН"])
    rr = ws3.max_row
    for c in range(1, 10):
        ws3.cell(rr, c).font = L1F; ws3.cell(rr, c).fill = L1P; ws3.cell(rr, c).border = THIN
        if c in (4,): ws3.cell(rr, c).number_format = N3
        if c in (5, 6, 7, 8): ws3.cell(rr, c).number_format = N2
    ws3.row_dimensions[rr].outlineLevel = 0
    for ln in sorted(r["накл"], key=lambda x: x["Дата"]):
        ws3.append(["    Накладна " + ln["Номер"], ln["Дата"], ln["Контрагент"], ln["Кіл"],
                    ln["СумаЗН"], ln["СумаДок"], ln["ОплаченоЗН"], ln["НеоплаченоЗН"],
                    "прихід"])
        rr = ws3.max_row
        for c in range(1, 10):
            ws3.cell(rr, c).font = INVF; ws3.cell(rr, c).border = THIN
            if c == 4: ws3.cell(rr, c).number_format = N3
            if c in (5, 6, 7, 8): ws3.cell(rr, c).number_format = N2
        if ln["НеоплаченоЗН"] > 0.005:
            ws3.cell(rr, 8).font = RED
        ws3.row_dimensions[rr].outlineLevel = 1
        for p in pays.get(ln["ДокКлюч"], []):
            ws3.append(["            " + p["Платіж"], p["ДатаПлатежу"], "", "", "", "",
                        round(p["Сума"] * ln["Частка"], 2), "", p["Тип"]])
            rr = ws3.max_row
            for c in range(1, 10):
                ws3.cell(rr, c).font = PAYF; ws3.cell(rr, c).border = THIN
                if c == 7: ws3.cell(rr, c).number_format = N2
            ws3.row_dimensions[rr].outlineLevel = 2
        if ln["НеоплаченоЗН"] > 0.005:
            ws3.append(["            ЗАЛИШОК НЕ СПЛАЧЕНО", "", "", "", "", "", "",
                        ln["НеоплаченоЗН"], "кредиторка"])
            rr = ws3.max_row
            for c in range(1, 10):
                ws3.cell(rr, c).font = RED; ws3.cell(rr, c).border = THIN
                if c == 8: ws3.cell(rr, c).number_format = N2
            ws3.row_dimensions[rr].outlineLevel = 2
ws3.freeze_panes = "A2"
ws3.sheet_properties.outlinePr.summaryBelow = False

# ---- 4. Неоплачені накладні ----
ws4 = wb.create_sheet("Неоплачені накладні")
head(ws4, ["Накладна", "Дата", "Постачальник", "Сума накладної, грн", "Борг, грн", "Позиції в накладній"],
     [16, 12, 34, 18, 16, 62])
seen = {}
for r in rows:
    for ln in r["накл"]:
        if ln["БоргДок"] > 0.005:
            e = seen.setdefault(ln["ДокКлюч"], {"Номер": ln["Номер"], "Дата": ln["Дата"],
                                                "Контрагент": ln["Контрагент"],
                                                "СумаДок": ln["СумаДок"], "Борг": ln["БоргДок"],
                                                "ЗН": []})
            e["ЗН"].append("%s (%s)" % (r["ЗН"], "{:,.2f}".format(ln["СумаЗН"])))
for e in sorted(seen.values(), key=lambda x: -x["Борг"]):
    ws4.append([e["Номер"], e["Дата"], e["Контрагент"], e["СумаДок"], e["Борг"], "; ".join(e["ЗН"])])
    rr = ws4.max_row
    for c in range(1, 7):
        ws4.cell(rr, c).border = THIN
        if c in (4, 5): ws4.cell(rr, c).number_format = N2
    ws4.cell(rr, 5).font = RED
ws4.append(["РАЗОМ", "", "", round(sum(e["СумаДок"] for e in seen.values()), 2),
            round(sum(e["Борг"] for e in seen.values()), 2), ""])
for c in range(1, 7):
    ws4.cell(ws4.max_row, c).font = Font(bold=True); ws4.cell(ws4.max_row, c).fill = TOTP
    if c in (4, 5): ws4.cell(ws4.max_row, c).number_format = N2
ws4.freeze_panes = "A2"; ws4.auto_filter.ref = "A1:F%d" % ws4.max_row

path = os.path.join(OUT, "МАТЕРИАЛИ_серпень_до_сплати.xlsx")
try:
    wb.save(path)
except PermissionError:
    path = os.path.join(OUT, "МАТЕРИАЛИ_серпень_до_сплати (нове).xlsx")
    wb.save(path)
    print("! файл був відкритий — збережено окремо")
print("\nOK ->", path)
print("аркуші:", wb.sheetnames)
json.dump({"rows": [{k: v for k, v in r.items() if k != "накл"} for r in rows], "T": T},
          open("data_august_pay.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
