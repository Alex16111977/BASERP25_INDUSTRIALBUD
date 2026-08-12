# -*- coding: utf-8 -*-
"""Excel-приложение к записке МД IRS 2026."""
import json, sys, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

OUT = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\А_ПланФактныйПроизводствоПолный\Письмо директору\04082026"
os.makedirs(OUT, exist_ok=True)

byON  = json.load(open("data_byon_calc.json", encoding="utf-8"))
mat   = json.load(open("data_materials.json", encoding="utf-8"))
match = json.load(open("data_match.json", encoding="utf-8"))
kazna = json.load(open("data_kazna.json", encoding="utf-8"))
budg  = json.load(open("data_budget.json", encoding="utf-8"))
xl    = json.load(open("data_excel_koshtoris.json", encoding="utf-8"))
audit = json.load(open("data_on_audit.json", encoding="utf-8"))
pos   = json.load(open("data_positions.json", encoding="utf-8"))["plan"]

HDR = PatternFill("solid", fgColor="1F4E79")
HF  = Font(color="FFFFFF", bold=True, size=10)
TOTF = Font(bold=True)
TOT = PatternFill("solid", fgColor="EEF3F8")
NUM = "#,##0.00"
NUM3 = "#,##0.###"

wb = openpyxl.Workbook()
wb.remove(wb.active)


def sheet(name, headers, rows, widths, numfmt=None, total_row=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR
        cell.font = HF
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.row_dimensions[1].height = 30
    for r in rows:
        ws.append(r)
    if numfmt:
        for col, fmt in numfmt.items():
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col).number_format = fmt
    if total_row:
        ws.append(total_row)
        for c in range(1, len(headers) + 1):
            ws.cell(ws.max_row, c).font = TOTF
            ws.cell(ws.max_row, c).fill = TOT
        if numfmt:
            for col, fmt in numfmt.items():
                ws.cell(ws.max_row, col).number_format = fmt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:" + get_column_letter(len(headers)) + str(ws.max_row)
    return ws


t = mat["tot"]
kz_mat = sum(x["Розхід"] for x in kazna if x["Стаття"] == "Строительные материалы")

# ---- 1. Зведення ----
ws = wb.create_sheet("Зведення")
summary = [
    ("МД IRS 2026 — зведення станом на 04.08.2026", None),
    ("", None),
    ("МАТЕРІАЛИ (грн з ПДВ)", None),
    ("План СС «15 м» (184 поз. × 6 будинків)", 4227672.06),
    ("План СС «30 м» (170 поз. × 7 будинків)", 7069053.18),
    ("План разом", t["ПланГрн"]),
    ("Факт закупівель (57 документів, липень 2026)", t["ФактГрн"]),
    ("Залишок до закупівлі (план на факт)", t["Осталось"]),
    ("Прогноз (факт + залишок)", t["Прогноз"]),
    ("Відхилення від плану (економія)", t["ОтклонениеПрогноз"]),
    ("Виконання плану закупівель, %", round(t["ФактГрн"] / t["ПланГрн"] * 100, 1)),
    ("", None),
    ("КОШТОРИС EXCEL (МАТЕРИАЛИ.xlsx)", None),
    ("IRS 15 — Σ всього", xl["15"]["СумаВсього"]),
    ("IRS 30 — Σ всього", xl["30"]["СумаВсього"]),
    ("Кошторис разом", round(xl["15"]["СумаВсього"] + xl["30"]["СумаВсього"], 2)),
    ("Розбіжність кошторис − СС", round(xl["15"]["СумаВсього"] + xl["30"]["СумаВсього"] - t["ПланГрн"], 2)),
    ("", None),
    ("ГРОШІ (Казна, регістр БДДС, липень 2026)", None),
    ("Надходження", round(sum(x["Прихід"] for x in kazna), 2)),
    ("Списання", round(sum(x["Розхід"] for x in kazna), 2)),
    ("у т.ч. Будівельні матеріали", round(kz_mat, 2)),
    ("", None),
    ("РОЗРИВ НАРАХУВАННЯ ↔ КАСА ПО МАТЕРІАЛАХ", None),
    ("Закуплено (прихід ТМЦ, ЕРП)", t["ФактГрн"]),
    ("Оплачено (Казна)", round(kz_mat, 2)),
    ("Неоплачена кредиторка", round(t["ФактГрн"] - kz_mat, 2)),
]
for a, b in summary:
    ws.append([a, b])
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 20
for r in range(1, ws.max_row + 1):
    ws.cell(r, 2).number_format = NUM
    v = ws.cell(r, 1).value or ""
    if v and ws.cell(r, 2).value is None and v == v.upper() and len(v) > 3:
        ws.cell(r, 1).font = Font(bold=True, color="1F4E79")
ws.cell(1, 1).font = Font(bold=True, size=13, color="1F4E79")

# ---- 2. Матеріали по загальних назвах ----
rows = []
for d in sorted(byON, key=lambda x: -x["ПланГрн"]):
    rows.append([d["ОН"], d["Этап"], d["Единица"], d["ПланКол"], d["ПланГрн"],
                 d["ФактКол"], d["ФактГрн"], d["Осталось"], d["Прогноз"], d["Відхилення"],
                 d["Процент"], d["ЦенаПлан"], d["ЦенаФакт"], d["План15"], d["План30"]])
sheet("Матеріали по ЗН",
      ["Загальна назва", "Етап", "Од.", "План кіл", "План грн", "Факт кіл", "Факт грн",
       "Залишок грн", "Прогноз грн", "Відхилення грн", "Факт/План %", "Ціна план",
       "Ціна факт", "План 15 м", "План 30 м"],
      rows, [34, 20, 7, 11, 14, 11, 14, 14, 14, 14, 10, 11, 11, 13, 13],
      {4: NUM3, 5: NUM, 6: NUM3, 7: NUM, 8: NUM, 9: NUM, 10: NUM, 11: "0.0",
       12: NUM, 13: NUM, 14: NUM, 15: NUM},
      ["РАЗОМ", "", "", round(sum(d["ПланКол"] for d in byON), 3), t["ПланГрн"],
       round(sum(d["ФактКол"] for d in byON), 3), t["ФактГрн"], t["Осталось"],
       t["Прогноз"], t["ОтклонениеПрогноз"], round(t["ФактГрн"] / t["ПланГрн"] * 100, 1),
       "", "", 4227672.06, 7069053.18])

# ---- 3. Ще закупити ----
rows = [[d["ОН"], d["Этап"], d["Единица"], d["ПланКол"], d["ФактКол"],
         round(max(d["ПланКол"] - d["ФактКол"], 0), 3), d["ПланГрн"], d["ФактГрн"],
         d["Осталось"], d["Процент"]]
        for d in sorted(byON, key=lambda x: -x["Осталось"]) if d["Осталось"] > 0]
sheet("Ще закупити",
      ["Загальна назва", "Етап", "Од.", "План кіл", "Факт кіл", "Треба кіл",
       "План грн", "Факт грн", "Залишок грн", "Факт/План %"],
      rows, [34, 20, 7, 11, 11, 11, 14, 14, 14, 11],
      {4: NUM3, 5: NUM3, 6: NUM3, 7: NUM, 8: NUM, 9: NUM, 10: "0.0"},
      ["РАЗОМ", "", "", "", "", "", round(sum(r[6] for r in rows), 2),
       round(sum(r[7] for r in rows), 2), round(sum(r[8] for r in rows), 2), ""])

# ---- 4. Кошторис vs СС ----
rows = []
for tag in ("15", "30"):
    m = match[tag]
    dom = m["Домов"]
    for d in m["розбіжності"]:
        rows.append(["IRS " + tag, d["Назва"], d["Ед"], d["КолExcel"], d["Кол1С"], d["ΔКол"],
                     d["ЦенаExcel"], d["Цена1С"], d["ΔЦена"], d["СумаExcel"], d["Сума1С"],
                     d["ΔСума"], round(d["ΔСума"] * dom, 2)])
    for e in m["тільки_в_Excel"]:
        rows.append(["IRS " + tag, e["Назва"], e["Ед"], e["КолДом"], 0, e["КолДом"],
                     e["Цена"], 0, e["Цена"], e["СумаДом"], 0, e["СумаДом"],
                     round(e["СумаДом"] * dom, 2)])
    for p in m["тільки_в_1С"]:
        rows.append(["IRS " + tag, p["ТекстСС"], p["Ед"], 0, p["КолДом"], -p["КолДом"],
                     0, p["Цена"], -p["Цена"], 0, p["СуммаДом"], -p["СуммаДом"],
                     round(-p["СуммаДом"] * dom, 2)])
sheet("Кошторис vs СС",
      ["Проєкт", "Позиція", "Од.", "Кіл Excel", "Кіл 1С", "Δ кіл", "Ціна Excel", "Ціна 1С",
       "Δ ціна", "Сума/дім Excel", "Сума/дім 1С", "Δ сума/дім", "Δ на всі будинки"],
      rows, [9, 42, 8, 11, 11, 10, 12, 12, 11, 15, 15, 14, 16],
      {4: NUM3, 5: NUM3, 6: NUM3, 7: NUM, 8: NUM, 9: NUM, 10: NUM, 11: NUM, 12: NUM, 13: NUM},
      ["РАЗОМ", "", "", "", "", "", "", "", "", "", "",
       round(sum(r[11] for r in rows), 2), round(sum(r[12] for r in rows), 2)])

# ---- 5. Графік закупівель ----
g15 = xl["15"]["график"]
g30 = xl["30"]["график"]
rows = [
    ["Липень", g15["липень"][1], g30["липень"][1],
     round(g15["липень"][1] + g30["липень"][1], 2), t["ФактГрн"]],
    ["Серпень", g15["серпень"][1], g30["серпень"][1],
     round(g15["серпень"][1] + g30["серпень"][1], 2), 0.0],
]
for r in rows:
    r.append(round(r[4] - r[3], 2))
sheet("Графік закупівель",
      ["Місяць", "План IRS 15", "План IRS 30", "План разом", "Факт закупівель", "Відхилення"],
      rows, [12, 16, 16, 16, 18, 16], {2: NUM, 3: NUM, 4: NUM, 5: NUM, 6: NUM},
      ["РАЗОМ", round(sum(r[1] for r in rows), 2), round(sum(r[2] for r in rows), 2),
       round(sum(r[3] for r in rows), 2), round(sum(r[4] for r in rows), 2),
       round(sum(r[5] for r in rows), 2)])

# ---- 6. Бюджет vs Каса ----
kz = {}
for x in kazna:
    kz[x["Стаття"]] = kz.get(x["Стаття"], 0.0) + x["Розхід"] - x["Прихід"]
rows = []
for b in budg:
    st = b["Стаття"]
    fk = kz.get(st)
    rows.append([st, b["Серпень"], b["Вересень"], b["Жовтень"], b["Бюджет"], b["Факт"],
                 round(fk, 2) if fk is not None else None,
                 round((fk if fk is not None else 0) - b["Факт"], 2)])
sheet("Бюджет vs Каса",
      ["Стаття ДДС", "План серпень", "План вересень", "План жовтень", "Бюджет (Excel)",
       "Факт (Excel)", "Факт Казна (липень)", "Δ Казна − Excel"],
      rows, [46, 15, 15, 15, 16, 15, 18, 16],
      {2: NUM, 3: NUM, 4: NUM, 5: NUM, 6: NUM, 7: NUM, 8: NUM})

# ---- 7. Аудит загальних назв ----
zero = [p for p in pos if p["СуммаДом"] == 0]
rows = []
for p in zero:
    on = p["ОН"] or ""
    of = audit["orph_fact"].get(on)
    rows.append(["Нульовий рядок кошторису", p["СС"][:24], p["ТекстСС"], on,
                 of["ФактГрн"] if of else 0.0,
                 "Закуплено, плану немає — дозаповнити кошторис" if of else "Закупівель не було"])
for on, noms in audit["nom_by_on"].items():
    if on == "АВР":
        for x in noms:
            if "введення резерву" not in x["Имя"]:
                rows.append(["Хибне ЗН", "", x["Код"] + " " + x["Имя"], on, x["Сумма"],
                             "Модульний автомат → ЗН «Вимикач» (план 124–142,50; АВР = 2 760)"])
    if on == "коліно":
        for x in noms:
            if "Profil" in x["Имя"]:
                rows.append(["Поза кошторисом", "", x["Код"] + " " + x["Имя"], on, x["Сумма"],
                             "Коліно ринви (система Profil), не каналізації — окремої позиції в кошторисі немає"])
for on, cap in (("шайба", "19,10 → факт 0,47"), ("Розподільча коробка", "100,47 → факт 24,14")):
    v = [d for d in byON if d["ОН"] == on]
    if v:
        rows.append(["Завищена планова ціна", "", on, on, v[0]["ПланГрн"],
                     "Планова ціна " + cap + "; призначення збігається — ЗН вірне, питання до кошторису"])
sheet("Аудит ЗН",
      ["Тип", "СС", "Позиція / картка", "Загальна назва", "Сума, грн", "Висновок"],
      rows, [26, 24, 52, 28, 14, 66], {5: NUM})

# ---- 8. Серпень: не купити зайве ----
ar = json.load(open("data_august_risk.json", encoding="utf-8"))
rows = [[r["ЗН"], r["ГрафікСерпень"], r["ПланГрн"], r["ФактГрн"], r["Викон%"],
         r["ЗалишокГрн"], r["Безпечно"], r["ЗайвеГрн"], r["Статус"], r["Позиції"]]
        for r in sorted(ar["rows"], key=lambda x: -x["ЗайвеГрн"])]
sheet("Серпень не купити зайве",
      ["Загальна назва", "Графік серпня", "План СС", "Вже куплено", "Викон. %",
       "Залишок потрібно", "Безпечно купити", "ЗАЙВЕ", "Статус", "Позиції кошторису"],
      rows, [32, 15, 14, 14, 10, 16, 16, 14, 24, 60],
      {2: NUM, 3: NUM, 4: NUM, 5: "0.0", 6: NUM, 7: NUM, 8: NUM},
      ["РАЗОМ", round(sum(r[1] for r in rows), 2), "", "", "",
       round(sum(r[5] for r in rows), 2), round(sum(r[6] for r in rows), 2),
       round(sum(r[7] for r in rows), 2), "", ""])

# ---- 9. Одиниці Excel vs СС ----
rows = [[u["Проєкт"], u["Позиція"], u["ЗН"], u["ЕдExcel"], u["Ед1С"], u["СумаСерпень"]]
        for u in sorted(ar["unit_bad"], key=lambda x: -x["СумаСерпень"])]
sheet("Одиниці Excel vs СС",
      ["Проєкт", "Позиція кошторису", "Загальна назва", "Од. Excel", "Од. 1С", "Сума серпня"],
      rows, [9, 46, 30, 12, 12, 15], {6: NUM},
      ["РАЗОМ", "", "", "", "", round(sum(r[5] for r in rows), 2)])

# ---- 10. Документи закупівлі ----
rows = [[d["Дата"], d["Номер"], d["Контрагент"], d["Сумма"]] for d in mat["docs"]]
sheet("Документи закупівлі", ["Дата", "Номер", "Контрагент", "Сума з ПДВ"],
      rows, [12, 18, 52, 16], {4: NUM},
      ["РАЗОМ", "", "", round(sum(r[3] for r in rows), 2)])

path = os.path.join(OUT, "Сверка_МД_IRS_2026_04-08-2026.xlsx")
try:
    wb.save(path)
except PermissionError:
    path = os.path.join(OUT, "Сверка_МД_IRS_2026_04-08-2026 (з серпнем).xlsx")
    wb.save(path)
    print("! основний файл відкритий в Excel — збережено окремо")
print("OK ->", path)
print("листов:", wb.sheetnames)
