# -*- coding: utf-8 -*-
"""Таблица 1:1 со структурой отчёта «План на факт за матеріалами»
(Підрозділ -> Загальна назва -> Аналітика -> Документ)
плюс две колонки из кошториса МАТЕРИАЛИ.xlsx: кіл-сть і сума серпня.

График серпня в Excel заведён по проектам, поэтому ложится на 4-й уровень
(у плановых строк Документ = карточка СС = проект).
"""
import win32com.client, sys, os, re, json, datetime
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

PODR = "МД IRS 2026"
OUT = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\А_ПланФактныйПроизводствоПолный\1С"
os.makedirs(OUT, exist_ok=True)

# ---------------- график серпня з кошториса ----------------
def norm(s):
    s = (s or "").lower().replace("ё", "е").replace("*", "х").replace("×", "х")
    s = s.replace(",", ".").replace("ʼ", "'").replace("’", "'")
    s = re.sub(r"[\s\u00a0]+", " ", s)
    return re.sub(r"[^0-9a-zа-яіїєґ'.х/]+", " ", s).strip()

xl = json.load(open("data_excel_koshtoris.json", encoding="utf-8"))
SSNAME = {"15": "МД IRS 2026 15 м (МД IRS 2026)", "30": "МД IRS 2026 30 м (МД IRS 2026)"}
aug = {}          # (ССname, norm(позиція)) -> [кіл, сума, од.Excel]
for tag, ss in SSNAME.items():
    for e in xl[tag]["rows"]:
        if e["СерпеньСума"] <= 0.004 and e["СерпеньКол"] <= 0.0004:
            continue
        k = (ss, norm(e["Назва"]))
        a = aug.setdefault(k, [0.0, 0.0, e["Ед"]])
        a[0] += e["СерпеньКол"]
        a[1] += e["СерпеньСума"]

# ---------------- дані звіту ----------------
v8 = win32com.client.Dispatch("V83.COMConnector")
erp = v8.Connect('Srvr="localhost";Ref="BaseERP";Usr="Администратор";Pwd="24043"')
S = erp.String
def f(v):
    try: return float(v)
    except: return 0.0

per = erp.NewObject("СтандартныйПериод")
per.ДатаНачала = datetime.datetime(2026, 1, 1, 0, 0, 0)
per.ДатаОкончания = datetime.datetime(2026, 12, 31, 23, 59, 59)
rep = erp.Отчеты.А_ПланФактныйПроизводствоПолный.Создать()
tz = rep.ПолучитьДанные(per)

# ЗН -> Аналітика -> Документ -> показники
tree = OrderedDict()
for i in range(tz.Количество()):
    r = tz.Получить(i)
    if S(r.Подразделение) != PODR or S(r.Блок) != "2. Матеріали":
        continue
    pk, fk, pg, fg = f(r.ПланКол), f(r.ФактКол), f(r.ПланГрн), f(r.ФактГрн)
    if pk == 0 and fk == 0 and pg == 0 and fg == 0:
        continue
    on = S(r.ОбщееНазвание) or "(без загальної назви)"
    an = S(r.Аналитика) or "(без аналітики)"
    dc = S(r.Документ) or ""
    ed = S(r.Единица)
    node = tree.setdefault(on, {"ед": ed, "an": OrderedDict()})
    if not node["ед"]:
        node["ед"] = ed
    a = node["an"].setdefault(an, {"ед": ed, "dc": OrderedDict()})
    d = a["dc"].setdefault(dc, [0.0, 0.0, 0.0, 0.0, 0.0, 0.0])  # пк,пг,фк,фг,авг_к,авг_с
    d[0] += pk; d[1] += pg; d[2] += fk; d[3] += fg

# августовский график -> плановые узлы уровня «Документ»
used = set()
for on, node in tree.items():
    for an, a in node["an"].items():
        for dc, d in a["dc"].items():
            k = (dc, norm(an))
            if k in aug and d[0] > 0:      # только плановые строки
                d[4] += aug[k][0]; d[5] += aug[k][1]
                used.add(k)
miss = {k: v for k, v in aug.items() if k not in used}
print("Позицій графіка серпня: %d, зіставлено: %d, не зіставлено: %d на %.2f грн"
      % (len(aug), len(used), len(miss), sum(v[1] for v in miss.values())))
for k, v in list(miss.items())[:10]:
    print("   не зіставлено:", k[1][:50], "%.2f" % v[1])

# ---------------- рендер ----------------
NUM3 = "#,##0.###"; NUM2 = "#,##0.00"
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "План на факт"

TITLE = Font(bold=True, size=14, color="1F4E79")
HF = Font(color="FFFFFF", bold=True, size=10)
HDR = PatternFill("solid", fgColor="1F4E79")
AUG_HDR = PatternFill("solid", fgColor="8A5A00")
L1F = Font(bold=True, size=11); L1P = PatternFill("solid", fgColor="BDD7EE")
L2F = Font(bold=True); L2P = PatternFill("solid", fgColor="DDEBF7")
L3F = Font(bold=False)
L4F = Font(italic=True, size=9, color="6B7683")
THIN = Side(style="thin", color="C9D2DA")

ws["A1"] = "План на факт за матеріалами"; ws["A1"].font = TITLE
ws["A2"] = "Параметри: Період 01.01.2026 – 31.12.2026;  Підрозділ у групі «МД IRS 2026»;  Блок «2. Матеріали»"
ws["A3"] = ("Колонки «Серпень» — графік закупівель з кошториса МАТЕРИАЛИ.xlsx (аркуші IRS 15 / IRS 30), "
            "проставлено на рядок проєкту. Увага: одиниці кошториса і СС по частині позицій різні.")
ws["A3"].font = Font(italic=True, size=9, color="8A5A00")

HEADERS = ["Підрозділ / Загальна назва / Аналітика / Документ", "Од.",
           "План, к-сть", "Ціна план, грн", "План, грн",
           "Факт, к-сть", "Ціна факт, грн", "Факт, грн", "Відхилення, к-сть",
           "Серпень, к-сть", "Серпень, сума грн"]
HROW = 5
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(HROW, c, h)
    cell.fill = AUG_HDR if c >= 10 else HDR
    cell.font = HF
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[HROW].height = 34
for c, w in enumerate([62, 8, 13, 14, 15, 13, 14, 15, 15, 13, 16], 1):
    ws.column_dimensions[get_column_letter(c)].width = w

def put(level, name, ed, pk, pg, fk, fg, ak, asum):
    r = ws.max_row + 1
    ws.cell(r, 1, ("    " * (level - 1)) + name)
    ws.cell(r, 2, ed)
    ws.cell(r, 3, pk if pk else None)
    ws.cell(r, 4, round(pg / pk, 2) if pk else None)
    ws.cell(r, 5, pg if pg else None)
    ws.cell(r, 6, fk if fk else None)
    ws.cell(r, 7, round(fg / fk, 2) if fk else None)
    ws.cell(r, 8, fg if fg else None)
    ws.cell(r, 9, round(fk - pk, 3) if (pk or fk) else None)
    ws.cell(r, 10, round(ak, 3) if ak else None)
    ws.cell(r, 11, round(asum, 2) if asum else None)
    fnt, fill = {1: (L1F, L1P), 2: (L2F, L2P), 3: (L3F, None), 4: (L4F, None)}[level]
    for c in range(1, 12):
        cc = ws.cell(r, c)
        cc.font = fnt
        if fill: cc.fill = fill
        cc.border = Border(bottom=THIN)
        cc.number_format = NUM3 if c in (3, 6, 9, 10) else (NUM2 if c in (4, 5, 7, 8, 11) else "General")
    ws.row_dimensions[r].outlineLevel = max(level - 1, 0)
    return r

def agg(dcs):
    t = [0.0] * 6
    for d in dcs:
        for j in range(6): t[j] += d[j]
    return t

tot = agg([d for n in tree.values() for a in n["an"].values() for d in a["dc"].values()])
put(1, PODR, "", tot[0], tot[1], tot[2], tot[3], tot[4], tot[5])

for on in sorted(tree, key=lambda k: -agg([d for a in tree[k]["an"].values() for d in a["dc"].values()])[1]):
    node = tree[on]
    t2 = agg([d for a in node["an"].values() for d in a["dc"].values()])
    put(2, on, node["ед"], *t2)
    for an in sorted(node["an"], key=lambda k: -agg(node["an"][k]["dc"].values())[1]):
        a = node["an"][an]
        t3 = agg(a["dc"].values())
        put(3, an, a["ед"], *t3)
        for dc in sorted(a["dc"], key=lambda k: -a["dc"][k][1]):
            d = a["dc"][dc]
            if dc:
                put(4, dc, "", *d)

ws.freeze_panes = "A%d" % (HROW + 1)
ws.auto_filter.ref = "A%d:K%d" % (HROW, ws.max_row)
ws.sheet_properties.outlinePr.summaryBelow = False

# ================= лист 2: зведення по ЗН + рішення по серпню =================
byON = {d["ОН"]: d for d in json.load(open("data_byon_calc.json", encoding="utf-8"))}
w2 = wb.create_sheet("Зведення по ЗН")
H2 = ["Загальна назва", "Од.", "План, к-сть", "Ціна план, грн", "План, грн",
      "Факт, к-сть", "Ціна факт, грн", "Факт, грн", "Відхилення, к-сть",
      "Серпень, к-сть", "Серпень, сума грн", "Залишок до закупівлі, грн",
      "Безпечно купити в серпні, грн", "ЗАЙВЕ в серпні, грн", "Рішення"]
w2.append(H2)
for c in range(1, len(H2) + 1):
    cell = w2.cell(1, c)
    cell.fill = AUG_HDR if 10 <= c <= 14 else HDR
    cell.font = HF
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
w2.row_dimensions[1].height = 40
for c, wd in enumerate([34, 8, 12, 14, 15, 12, 14, 15, 14, 12, 15, 17, 18, 15, 24], 1):
    w2.column_dimensions[get_column_letter(c)].width = wd

RED = Font(color="B3261E", bold=True)
AMB = Font(color="8A5A00")
rows2 = []
for on, node in tree.items():
    t = agg([d for a in node["an"].values() for d in a["dc"].values()])
    b = byON.get(on, {})
    ost = b.get("Осталось", 0.0)
    graf = t[5]
    over = round(max(graf - ost, 0.0), 2) if graf else 0.0
    if graf <= 0.004:
        dec = "немає у графіку серпня"
    elif over > 0.005 and ost <= 0.005:
        dec = "НЕ КУПУВАТИ — закрито"
    elif over > 0.005:
        dec = "зменшити до залишку"
    else:
        dec = "купувати за графіком"
    rows2.append([on, node["ед"], t[0], round(t[1] / t[0], 2) if t[0] else None, t[1],
                  t[2], round(t[3] / t[2], 2) if t[2] else None, t[3],
                  round(t[2] - t[0], 3), t[4] or None, graf or None, ost,
                  round(min(graf, ost), 2) if graf else None, over or None, dec])
rows2.sort(key=lambda r: (-(r[13] or 0), -(r[10] or 0)))
for r in rows2:
    w2.append(r)
    rr = w2.max_row
    for c in range(1, 16):
        cc = w2.cell(rr, c)
        cc.border = Border(bottom=THIN)
        cc.number_format = NUM3 if c in (3, 6, 9, 10) else (NUM2 if c in (4, 5, 7, 8, 11, 12, 13, 14) else "General")
    if r[14].startswith("НЕ КУПУВАТИ"):
        for c in range(1, 16): w2.cell(rr, c).font = RED
    elif r[14].startswith("зменшити"):
        for c in range(1, 16): w2.cell(rr, c).font = AMB
tt = agg([d for n in tree.values() for a in n["an"].values() for d in a["dc"].values()])
w2.append(["РАЗОМ", "", tt[0], None, tt[1], tt[2], None, tt[3], round(tt[2] - tt[0], 3),
           None, tt[5], round(sum(r[11] or 0 for r in rows2), 2),
           round(sum(r[12] or 0 for r in rows2), 2), round(sum(r[13] or 0 for r in rows2), 2), ""])
for c in range(1, 16):
    w2.cell(w2.max_row, c).font = Font(bold=True)
    w2.cell(w2.max_row, c).fill = PatternFill("solid", fgColor="EEF3F8")
    w2.cell(w2.max_row, c).number_format = NUM3 if c in (3, 6, 9, 10) else (NUM2 if c in (4, 5, 7, 8, 11, 12, 13, 14) else "General")
w2.freeze_panes = "A2"
w2.auto_filter.ref = "A1:O%d" % w2.max_row

# ================= лист 3: сім закритих позицій =================
w3 = wb.create_sheet("Не купувати в серпні")
w3.append(H2[:14] + ["Чому закрито"])
for c in range(1, 16):
    cell = w3.cell(1, c)
    cell.fill = AUG_HDR if 10 <= c <= 14 else HDR
    cell.font = HF
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
w3.row_dimensions[1].height = 40
for c, wd in enumerate([34, 8, 12, 14, 15, 12, 14, 15, 14, 12, 15, 17, 18, 15, 52], 1):
    w3.column_dimensions[get_column_letter(c)].width = wd
for r in rows2:
    if not r[14].startswith("НЕ КУПУВАТИ"):
        continue
    pk, fk = r[2], r[5]
    pct = (fk / pk * 100) if pk else 0
    why = "куплено %s з %s %s — %.1f%% кількості" % (
        ("{:,.3f}".format(fk)).rstrip("0").rstrip("."),
        ("{:,.3f}".format(pk)).rstrip("0").rstrip("."), r[1] or "од.", pct)
    if fk > pk:
        why = "ПЕРЕКУПЛЕНО: " + why
    w3.append(r[:14] + [why])
    rr = w3.max_row
    for c in range(1, 16):
        cc = w3.cell(rr, c)
        cc.border = Border(bottom=THIN)
        cc.font = RED if c == 15 or r[13] else Font()
        cc.number_format = NUM3 if c in (3, 6, 9, 10) else (NUM2 if c in (4, 5, 7, 8, 11, 12, 13, 14) else "General")
closed = [r for r in rows2 if r[14].startswith("НЕ КУПУВАТИ")]
w3.append(["РАЗОМ", "", "", "", "", "", "", "", "", None,
           round(sum(r[10] or 0 for r in closed), 2), round(sum(r[11] or 0 for r in closed), 2),
           None, round(sum(r[13] or 0 for r in closed), 2), ""])
for c in range(1, 16):
    w3.cell(w3.max_row, c).font = Font(bold=True)
    w3.cell(w3.max_row, c).fill = PatternFill("solid", fgColor="EEF3F8")
    w3.cell(w3.max_row, c).number_format = NUM2
w3.freeze_panes = "A2"

path = os.path.join(OUT, "План на факт з графіком серпня.xlsx")
try:
    wb.save(path)
except PermissionError:
    path = os.path.join(OUT, "План на факт з графіком серпня (нове).xlsx")
    wb.save(path)
    print("! файл був відкритий — збережено окремо")
print("OK ->", path)
print("рядків:", ws.max_row, "| ЗН:", len(tree))
print("КОНТРОЛЬ: План=%.2f Факт=%.2f Серпень=%.2f" % (tot[1], tot[3], tot[5]))
