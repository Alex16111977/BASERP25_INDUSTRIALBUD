# -*- coding: utf-8 -*-
"""Проба структуры файла «ЗВіт  виконання IRS 15 м.кв. 06.08.26.xlsx».

Регламент/LESSONS: макет книги у каждого прораба свой — каждый новый файл
смотреть глазами ДО запуска загрузчика (листы, заголовки, колонки A/B, колонка факта).
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

XL = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Загрузка СС\06082026\ЗВіт  виконання IRS 15 м.кв. 06.08.26.xlsx"

wb = openpyxl.load_workbook(XL, data_only=True)
print("Листы:", wb.sheetnames)

for имя in wb.sheetnames:
    ws = wb[имя]
    print(f"\n=== Лист «{имя}»: строк {ws.max_row}, колонок {ws.max_column} ===")

    # строка 1 — заголовки
    for c in range(1, min(ws.max_column, 15) + 1):
        v = ws.cell(1, c).value
        if v is not None:
            print(f"  R1C{c}: {str(v)[:100]!r}")

    # колонка факта: ищем по ослабленному ключу «виконання робіт» (урок IRS 30 м)
    колф = None
    for c in range(1, min(ws.max_column, 15) + 1):
        if "виконання робіт" in str(ws.cell(1, c).value or "").lower():
            колф = c
            break
    print(f"  Колонка факта по заголовку: {колф}")

    # все строки: A, B и значение факта с number_format
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        f = ws.cell(r, колф).value if колф else None
        nf = ws.cell(r, колф).number_format if колф else ""
        if a is None and b is None and f is None:
            continue
        print(f"  R{r}: A={str(a)[:48]!r} B={str(b)[:28]!r} F={f!r} [{nf}]")

wb.close()
print("\nПроба завершена.")
