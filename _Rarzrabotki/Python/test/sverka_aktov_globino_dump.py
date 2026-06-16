# -*- coding: utf-8 -*-
"""Дамп листа 'Сверка' результата для анализа расхождений."""
import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

f = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры\Сверка актов Глобино.xlsx"
wb = openpyxl.load_workbook(f, data_only=True)
ws = wb["Сверка"]
print(f"{'Номер':<13}{'ЕРП':>14}{'Бух205':>14}{'Бух2051':>14}{'БухИтог':>14}{'Разница':>12}  Статус / Док Бух")
for row in ws.iter_rows(min_row=2, values_only=True):
    (_, num, date, erp_doc, erp_sum, buh_doc, b205, b2051, bsum, diff, status) = row
    if num is None:
        continue
    def fmt(v):
        return f"{v:>14,.2f}" if isinstance(v, (int, float)) else " " * 14
    d = f"{diff:>12,.2f}" if isinstance(diff, (int, float)) else " " * 12
    print(f"{num:<13}{fmt(erp_sum)}{fmt(b205)}{fmt(b2051)}{fmt(bsum)}{d}  {status} | {buh_doc}")
print()
ws3 = wb["Контроль"]
for row in ws3.iter_rows(min_row=2, values_only=True):
    n, v = row
    if isinstance(v, float):
        print(f"  {n:<62}{v:>18,.2f}")
    else:
        print(f"  {n:<62}{v}")
