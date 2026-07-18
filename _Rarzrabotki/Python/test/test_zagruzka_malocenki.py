# -*- coding: utf-8 -*-
"""Smoke-тест обработки «Загрузка ПередачаМалоценныхАктивовВЭксплуатацию» (BuhBud).

1. СведенияОВнешнейОбработке() — БСП-регистрация (Вид, Назначение).
2. Тестовый документ ПередачаМалоценныхАктивовВЭксплуатацию: get-or-create по
   Комментарию-маркеру (НЕ проводится, НЕ удаляется).
3. Таблица кодов из тестового xlsx (повтор клиентской логики чтения:
   ПерваяСтрока=20, КолонкаКод=7, КолонкаКоличество=10, скан до 50 пустых).
4. обр.ЗаполнитьМалоценныеАктивы(ДокСсылка, Таблица) → сверка строк ТЧ,
   суммы Количество, заполненности счетов.

Поведение как у образца: функция ДОБАВЛЯЕТ строки к ТЧ без очистки,
поэтому тест сам чистит ТЧ документа перед вызовом.
"""
import sys, os
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import win32com.client

EPF = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Обработки\Загрузка ПередачаМалоценныхАктивовВЭксплуатацию.epf"
XLSX = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Загрузка в документ ПередачаМалоценныхАктивовВЭксплуатацию\2\++М29 АВР №22_1.10_Ар.xlsx"
SHEET = "М29 №21 1.10 - АР"
FIRST_ROW, COL_CODE, COL_QTY = 20, 7, 10
MARKER = "ТЕСТ ЗагрузкаМалоценки (Claude) — не проводить"
# коды с остатком на счетах малоценки (проверено ostatki_discovery): Пензель плоский -> 221
EXPECTED_WITH_ACC = {"00000004876": "221"}

fails = []

def check(cond, msg):
    print(("  OK: " if cond else "  FAIL: ") + msg)
    if not cond:
        fails.append(msg)

# --- Excel: повтор клиентской логики образца ---
def read_codes():
    ws = openpyxl.load_workbook(XLSX, data_only=True)[SHEET]
    rows, blanks, r = [], 0, FIRST_ROW
    while blanks < 50:
        raw = ws.cell(row=r, column=COL_CODE).value
        if raw is None or str(raw).strip() == "":
            blanks += 1
            r += 1
            continue
        blanks = 0
        code = raw.strip() if isinstance(raw, str) else str(int(raw)).rjust(11, "0")
        try:
            qty = float(ws.cell(row=r, column=COL_QTY).value)
        except (TypeError, ValueError):
            qty = 0
        if qty > 0:
            rows.append((code, qty))
        r += 1
    return rows

codes = read_codes()
total_qty = round(sum(q for _, q in codes), 3)
print(f"Excel '{SHEET}': строк с кодом и qty>0: {len(codes)}, sum(qty)={total_qty}")

# --- COM ---
v8 = win32com.client.Dispatch("V83.COMConnector")
buh = v8.Connect('Srvr="localhost";Ref="bas_industrialbud";Usr="cfo";Pwd="2442"')
S = buh.String

print("\n[1] СведенияОВнешнейОбработке()")
obr = buh.ВнешниеОбработки.Создать(EPF, False)
sved = obr.СведенияОВнешнейОбработке()
naznach = sved.Назначение
check(S(sved.Вид) == "ЗаполнениеОбъекта", f"Вид = {S(sved.Вид)}")
check(naznach.Количество() == 1 and S(naznach.Получить(0)) == "Документ.ПередачаМалоценныхАктивовВЭксплуатацию",
      f"Назначение = {S(naznach.Получить(0))}")
check("малоценных активов" in S(sved.Наименование), f"Наименование = {S(sved.Наименование)}")

print("\n[2] Тестовый документ (get-or-create, не проводить, не удалять)")
q = buh.NewObject("Запрос")
q.Text = """
ВЫБРАТЬ ПЕРВЫЕ 1 Док.Ссылка КАК Ссылка
ИЗ Документ.ПередачаМалоценныхАктивовВЭксплуатацию КАК Док
ГДЕ ВЫРАЗИТЬ(Док.Комментарий КАК Строка(200)) = &Коммент И НЕ Док.Проведен
"""
q.SetParameter("Коммент", MARKER)
r = q.Execute().Выгрузить()
if r.Количество():
    dok = r.Получить(0).Ссылка.ПолучитьОбъект()
    print("  найден:", S(dok.Номер), S(dok.Дата))
else:
    import datetime
    dok = buh.Документы.ПередачаМалоценныхАктивовВЭксплуатацию.СоздатьДокумент()
    dok.Дата = datetime.datetime.now().replace(microsecond=0)
    dok.Комментарий = MARKER
    oq = buh.NewObject("Запрос")
    oq.Text = "ВЫБРАТЬ ПЕРВЫЕ 1 Орг.Ссылка КАК Ссылка ИЗ Справочник.Организации КАК Орг ГДЕ НЕ Орг.ПометкаУдаления"
    dok.Организация = oq.Execute().Выгрузить().Получить(0).Ссылка
    sq = buh.NewObject("Запрос")
    sq.Text = "ВЫБРАТЬ ПЕРВЫЕ 1 Скл.Ссылка КАК Ссылка ИЗ Справочник.Склады КАК Скл ГДЕ НЕ Скл.ПометкаУдаления И НЕ Скл.ЭтоГруппа"
    dok.Склад = sq.Execute().Выгрузить().Получить(0).Ссылка
    dok.Записать()
    print("  создан:", S(dok.Номер))

# тест сам обеспечивает чистую ТЧ (функция по дизайну добавляет без очистки)
dok.МалоценныеАктивы.Очистить()
dok.Записать()
dok_ref = dok.Ссылка

print("\n[3] ЗаполнитьМалоценныеАктивы(ДокСсылка, Таблица)")
tz = buh.NewObject("ТаблицаЗначений")
tz.Колонки.Добавить("КодНоменлатуры")
tz.Колонки.Добавить("КвоДляСписания")
for code, qty in codes:
    row = tz.Добавить()
    row.КодНоменлатуры = code
    row.КвоДляСписания = qty

res = obr.ЗаполнитьМалоценныеАктивы(dok_ref, tz)
print(f"  протокол: Загружено={res.Загружено}, НеНайденоКодов={res.НеНайденоКодов}, "
      f"БезСчета={res.БезСчета}, Ошибка={res.Ошибка}, ТекстОшибки='{S(res.ТекстОшибки)}'")
check(not res.Ошибка, "Ошибка = Ложь")
check(res.Загружено == len(codes), f"Загружено {res.Загружено} == строк Excel {len(codes)}")
check(res.НеНайденоКодов == 0, f"НеНайденоКодов = {res.НеНайденоКодов}")

print("\n[4] Сверка документа после Записать()")
dok2 = dok_ref.ПолучитьОбъект()
tch = dok2.МалоценныеАктивы
check(tch.Количество() == len(codes), f"строк ТЧ {tch.Количество()} == {len(codes)}")
sum_qty = round(sum(tch.Получить(i).Количество for i in range(tch.Количество())), 3)
check(sum_qty == total_qty, f"Σ Количество ТЧ {sum_qty} == Excel {total_qty}")

with_acc, empty_acc = {}, 0
for i in range(tch.Количество()):
    row = tch.Получить(i)
    kod_nom = S(row.Номенклатура.Код)
    kod_acc = S(row.СчетУчетаБУ.Код) if S(row.СчетУчетаБУ) else ""
    if S(row.ЕдиницаИзмерения) == "" or row.Коэффициент != 1:
        fails.append(f"{kod_nom}: ЕдИзм='{S(row.ЕдиницаИзмерения)}', Коэф={row.Коэффициент}")
    if kod_acc:
        with_acc[kod_nom] = kod_acc
    else:
        empty_acc += 1
print(f"  со счётом: {len(with_acc)} {with_acc}, без счёта: {empty_acc}")
for code, acc in EXPECTED_WITH_ACC.items():
    check(with_acc.get(code) == acc, f"{code} -> счёт {with_acc.get(code)} (ожидали {acc})")
check(empty_acc == res.БезСчета, f"пустых счетов в ТЧ {empty_acc} == БезСчета {res.БезСчета}")
check(not dok2.Проведен, "документ НЕ проведён")

print("\n[5] Идемпотентность (повторный вызов ДОБАВЛЯЕТ строки — поведение образца)")
res2 = obr.ЗаполнитьМалоценныеАктивы(dok_ref, tz)
dok3 = dok_ref.ПолучитьОбъект()
check(dok3.МалоценныеАктивы.Количество() == 2 * len(codes),
      f"после 2-го вызова строк {dok3.МалоценныеАктивы.Количество()} == {2*len(codes)} (добавление без очистки)")
# вернуть один комплект строк
dok3.МалоценныеАктивы.Очистить()
dok3.Записать()
res3 = obr.ЗаполнитьМалоценныеАктивы(dok_ref, tz)

print("\n" + ("=== SMOKE OK ===" if not fails else f"=== SMOKE FAIL: {len(fails)} ==="))
for f in fails:
    print("  -", f)
sys.exit(1 if fails else 0)
