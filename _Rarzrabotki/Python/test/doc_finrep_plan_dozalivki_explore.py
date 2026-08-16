# -*- coding: utf-8 -*-
"""Разведка Excel для плана дозаливки: секции, строки, суммы. ТОЛЬКО ЧТЕНИЕ Excel."""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

EXCEL = r"C:/Configuration_downloads/BASERP25/_Rarzrabotki/notebook/knowledge_А_ФинансовыйОтчетПроизводства/Отчеты/Финансовый_отчет Производства  14-08-2026.xlsx"
ЛИСТ = "24,07,2026"

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb[ЛИСТ]
print(f"max_row={ws.max_row}, max_col={ws.max_column}")

def число(значение):
    if isinstance(значение, (int, float)):
        return float(значение)
    if isinstance(значение, str):
        try:
            return float(значение.replace("\xa0", "").replace(" ", "").replace(",", "."))
        except ValueError:
            return None
    return None

for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1)
    b = ws.cell(r, 2)
    c = ws.cell(r, 3)
    d = ws.cell(r, 4)
    имя = " ".join(str(a.value).split()) if a.value is not None else ""
    прим = " ".join(str(b.value).split()) if b.value is not None else ""
    if not имя and not прим and c.value is None and d.value is None:
        continue
    метка = "BOLD" if (a.value is not None and a.font.bold) else "    "
    cс = число(c.value)
    dс = число(d.value)
    cт = f"{cс:,.2f}" if cс is not None else (str(c.value) if c.value is not None else "")
    dт = f"{dс:,.2f}" if dс is not None else (str(d.value) if d.value is not None else "")
    print(f"r{r:3d} {метка} A={имя[:55]:55} | B={прим[:35]:35} | C={cт:>15} | D={dт:>15}")
wb.close()
