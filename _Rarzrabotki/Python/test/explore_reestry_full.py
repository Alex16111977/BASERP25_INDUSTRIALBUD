# -*- coding: utf-8 -*-
"""Полные значения: лист ЕРП целиком + примеры колонки Документ из Бухгалтерии."""
import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

ERP = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры\Строительные материалы ЕРП.xlsx"
BUH = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры\Строительные материалы из Бухгалтерии.xlsx"

print("=" * 100)
print("ЕРП / Лист2 — полностью:")
wb = openpyxl.load_workbook(ERP, read_only=True, data_only=True)
ws = wb["Лист2"]
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    vals = [("" if v is None else str(v).replace("\n", "\\n")) for v in row]
    # обрезаем хвостовые пустые
    while vals and vals[-1] == "":
        vals.pop()
    print(f"  r{i:>3}: " + " || ".join(vals))
wb.close()

print("=" * 100)
print("БУХ — уникальные значения 'Документ' (Глобино-2 и Глобино-2051):")
wb = openpyxl.load_workbook(BUH, read_only=True, data_only=True)
for sheet in ["Глобино-2", "Глобино-2051"]:
    ws = wb[sheet]
    docs = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        doc = row[1]
        if doc:
            docs[str(doc).replace("\n", "\\n")] = docs.get(str(doc).replace("\n", "\\n"), 0) + 1
    print("-" * 100)
    print(f"  SHEET {sheet}: уникальных документов = {len(docs)}")
    for j, (d, cnt) in enumerate(sorted(docs.items())):
        print(f"    [{cnt:>3} строк] {d}")
        if j >= 40:
            print(f"    ... и ещё {len(docs) - 41}")
            break
wb.close()
print("DONE")
