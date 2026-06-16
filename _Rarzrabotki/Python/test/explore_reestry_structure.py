# -*- coding: utf-8 -*-
"""Разведка структуры двух файлов реестров стройматериалов."""
import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

FILES = [
    r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры\Строительные материалы из Бухгалтерии.xlsx",
    r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Реестры\Строительные материалы ЕРП.xlsx",
]

for path in FILES:
    print("=" * 100)
    print("FILE:", path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print("  ERROR:", e)
        continue
    print("  Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print("-" * 100)
        print(f"  SHEET: {sheet_name}  max_row={ws.max_row} max_col={ws.max_column}")
        # первые 15 строк, до 12 колонок
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, max_col=12, values_only=True), start=1):
            vals = []
            for v in row:
                s = str(v) if v is not None else ""
                if len(s) > 30:
                    s = s[:27] + "..."
                vals.append(s)
            print(f"    r{i:>3}: " + " | ".join(vals))
    wb.close()
print("DONE")
