# -*- coding: utf-8 -*-
# Сверка снимка 24/25.07 (файл) против сохранённого 17.07 (база): регрессии и полная матрица.
import win32com.client, sys, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

F24 = r"C:\Configuration_downloads\BASERP25\_Rarzrabotki\Загрузка СС\Виробництво\ЗВіт  виконання IRS 15 м.кв. 24.07.26.xlsx"
N = 6
def норм(s): return re.sub(r"\s+"," ",str(s or "").strip().lower().replace("'","").replace("’",""))
def парс(t,n):
    t=str(t or "").strip()
    if not t: return None
    s=re.findall(r"(\d+)\s*модул\w*\s*-*\s*(\d+(?:[.,]\d+)?)\s*%",t,re.I)
    if s:
        o=[]
        for k,p in s: o+=[float(p.replace(",","."))]*int(k)
        return (o+[0.0]*n)[:n]
    m=re.fullmatch(r"(\d+(?:[.,]\d+)?)\s*%",t)
    return [float(m.group(1).replace(",","."))]*n if m else None
def факт(v,nf):
    if v is None or isinstance(v,bool): return ""
    if isinstance(v,(int,float)): return (f"{v*100:g}%" if "%" in (nf or "") else f"{v:g}")
    return str(v).strip()

wb=openpyxl.load_workbook(F24,data_only=True)
имя=[n for n in wb.sheetnames if n.strip().lower().startswith("звіт") and n.strip().lower()!="звіт"][0]
ws=wb[имя]; колф=8
for c in range(1,16):
    if "факт виконання робіт" in str(ws.cell(1,c).value or "").lower(): колф=c; break
new={}
order=[]
for r in range(2,ws.max_row+1):
    nz=str(ws.cell(r,1).value or "").strip()
    if not nz or "Загальна" in nz or nz.lower()=="табель": continue
    p=парс(факт(ws.cell(r,колф).value, ws.cell(r,колф).number_format), N)
    new[норм(nz)]=(nz, p if p is not None else [0.0]*N)
    order.append(норм(nz))

erp=win32com.client.Dispatch("V83.COMConnector").Connect('Srvr="localhost";Ref="BaseERP";Usr="Администратор";Pwd="24043"')
сс=erp.Справочники.А_СтруктураСебестоимости.НайтиПоНаименованию("МД IRS 2026 15 м (МД IRS 2026)")
# сохранённый 17.07: % по домику×работе из ТЧ документов дня 17
q=erp.NewObject("Запрос"); q.УстановитьПараметр("СС",сс.Ссылка)
q.Текст=("ВЫБРАТЬ Д.Подразделение.Наименование КАК Подр, Т.Работа.Наименование КАК Раб, Т.Процент КАК П "
         "ИЗ Документ.А_ВыполнениеРаботМодулей.Работы КАК Т "
         "ЛЕВОЕ СОЕДИНЕНИЕ Документ.А_ВыполнениеРаботМодулей КАК Д ПО Д.Ссылка=Т.Ссылка "
         "ГДЕ Д.СтруктураСебестоимости=&СС И ДЕНЬ(Д.Дата)=17")
т=q.Execute().Выгрузить()
old={}
for i in range(т.Количество()):
    s=т.Получить(i); m=re.search(r"№(\d+)",str(s.Подр))
    if m: old.setdefault(норм(s.Раб),{})[int(m.group(1))]=float(s.П)

print(f"Лист: «{имя}» -> дата 25.07.2026 (в имени файла 24.07); колонка факта {колф}\n")
print("РЕГРЕССИИ (24.07 < 17.07 хоть по одному домику):")
рег=0
for nk in order:
    nz, pnew = new[nk]
    po = old.get(nk, {})
    паден=[]
    for k in range(1,N+1):
        o=po.get(k,0.0); n_=pnew[k-1]
        if n_ < o - 0.001: паден.append(f"№{k}:{int(o)}→{int(n_)}")
    if паден:
        рег+=1
        print(f"   ⚠ {nz[:42]:42} " + ", ".join(паден))
if not рег: print("   нет — все работы не убывают ✓")

print("\nПОЛНАЯ МАТРИЦА 24/25.07 (проценты в М1 по домикам):")
print(f"   {'Работа':44} №1 №2 №3 №4 №5 №6")
for nk in order:
    nz,p=new[nk]
    print(f"   {nz[:44]:44} " + " ".join(f"{int(p[k]):>2}" for k in range(N)))
