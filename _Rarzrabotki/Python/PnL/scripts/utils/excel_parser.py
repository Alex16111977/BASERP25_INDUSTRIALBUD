"""Parser for PL Excel sheets."""
import datetime
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
import config
try:
    from . import fake_articles
except ImportError:  # запуск вне пакета utils
    import fake_articles


# Признак строки-комментария финансиста: число + единица валюты в названии
# (напр. "Кернел кухня - - 4 554 246грн", "ВООЗ 2025 - 3,7 млн грн согл сс").
# Такие строки НЕ являются статьями каталога — это поясняющие подытоги/комментарии
# из листов Excel, которые финансист пишет в столбце B.
_COMMENT_VALUE_RE = re.compile(
    r"\d[\d\s ,.\-]*\s*(грн|тыс|млн|тис)\b",
    re.IGNORECASE,
)


def _col_idx(letter):
    return column_index_from_string(letter) - 1


def _parse_month_key(val):
    """Return 'YYYY-MM-01' if val is a month-start date; else None."""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-01")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-01")
    return None


def find_month_columns(ws, month_header_str):
    """Scan row 2 for the month marker, return (ci_art, ci_f1, ci_f2, ci_total, ci_pct, ci_cmt).
    Article always in B; amount columns come right after the month marker cell.
    """
    ci_art = _col_idx("B")
    target = str(month_header_str)[:10]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        key = _parse_month_key(v)
        if key == target:
            # Excel column index is 1-based; convert to 0-based
            base = c - 1
            return ci_art, base, base + 1, base + 2, base + 3, base + 4
    return None


def is_summary_sheet(name: str) -> bool:
    if name in config.SUMMARY_SHEET_EXACT:
        return True
    for p in config.SUMMARY_SHEET_PATTERNS:
        if name.startswith(p):
            return True
    return False


def _norm_name(s):
    return " ".join(str(s).split()).strip().casefold()


_GROUP_NAMES_LC = {_norm_name(g) for g in config.PL_GROUP_NAMES}
_REPORT_END_LC = {_norm_name(m) for m in getattr(config, "REPORT_END_MARKERS", [])}


def classify_row(article: str) -> str:
    """Returns: 'group' (grouping header) | 'summary' (total/%) | 'detail'."""
    if not article:
        return "summary"
    s = str(article).strip()
    if not s:
        return "summary"
    s_norm = _norm_name(s)
    if s_norm in _GROUP_NAMES_LC:
        return "group"
    for kw in config.TOTAL_ROW_KEYWORDS:
        if kw.lower() in s.lower():
            return "summary"
    # Комментарий финансиста с числом и единицей валюты в названии — не статья.
    if _COMMENT_VALUE_RE.search(s):
        return "summary"
    # Явные подгруппы/комментарии финансиста (точное имя без чисел).
    if s_norm in {_norm_name(x) for x in getattr(config, "EXTRA_COMMENT_NAMES", set())}:
        return "summary"
    # Реестр фейк-статей (data/fake_articles.json) — служебные строки финансиста по имени.
    if fake_articles.match_fake(s):
        return "summary"
    return "detail"


def is_total_row(article: str) -> bool:
    return classify_row(article) != "detail"


def _to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_workbook(xlsx_path: str, month_header: str):
    """Yields dicts: {sheet_name, rows: [{row, article, sum_f1, sum_f2, total, percent, comment}]}

    month_header — "YYYY-MM-01" marker used to locate factual month columns in row 2.
    """
    wb = load_workbook(xlsx_path, data_only=True)

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        name = ws.title
        # is_summary_sheet фильтр НЕ применяем:
        # все visible PnL-листы (включая сводные PL_ЦО, PL_Логистика, Техника, MAN №*) —
        # это полноценные PnL-листы с данными по статьям, которые финансист хочет видеть
        # в 1С как отдельные документы А_ОтчетPL. Маппинг лист→подразделение+направление
        # задаёт документ А_РасшифровкаЛистов (источник истины — UI 1С).
        # Листы без PnL-формата (Метрики, Индекс, Настройки) отсекаются ниже через find_month_columns.
        cols = find_month_columns(ws, month_header)
        if cols is None:
            continue
        ci_art, ci_f1, ci_f2, ci_tot, ci_pct, ci_cmt = cols
        max_col = max(ci_art, ci_f1, ci_f2, ci_tot, ci_pct, ci_cmt) + 1

        rows_out = []
        group_totals = []  # строки-подытоги групп (Excel, как есть)
        general_totals = []  # общие итоги низа листа (Маржинальный доход и т.п.)
        current_group = ""
        report_ended = False  # стали ли мы ниже финальных итогов листа (рентабельность/прибыль в распоряжении)
        for r_idx in range(config.PL_DATA_START_ROW, ws.max_row + 1):
            article = ws.cell(r_idx, ci_art + 1).value
            if article is None:
                continue
            s = str(article).strip()
            if not s:
                continue

            sum_f1 = _to_number(ws.cell(r_idx, ci_f1 + 1).value)
            sum_f2 = _to_number(ws.cell(r_idx, ci_f2 + 1).value)
            total = _to_number(ws.cell(r_idx, ci_tot + 1).value)
            comment_raw = ws.cell(r_idx, ci_cmt + 1).value
            comment = str(comment_raw).strip() if comment_raw is not None else ""

            kind = classify_row(s)

            # Группа-подытог: и запоминаем как "текущую группу", и сохраняем её сумму
            if kind == "group":
                current_group = s
                if sum_f1 is not None or sum_f2 is not None or total is not None:
                    group_totals.append({
                        "row": r_idx,
                        "group_name": s,
                        "sum_f1_excel": sum_f1 or 0.0,
                        "sum_f2_excel": sum_f2 or 0.0,
                        "sum_excel": total or 0.0,
                    })
                continue

            # Общие итоги низа листа → ТЧ ИтогиОбщие
            if kind == "summary":
                vid = config.classify_summary(s)
                if vid and (sum_f1 is not None or sum_f2 is not None or total is not None):
                    general_totals.append({
                        "row": r_idx,
                        "show": s,
                        "vid": vid,
                        "sum_f1": sum_f1 or 0.0,
                        "sum_f2": sum_f2 or 0.0,
                        "sum_total": total or 0.0,
                    })
                # Структурный фильтр фейков: после финальных итогов (рентабельность /
                # прибыль в распоряжении) detail-строки ниже — служебные заметки финансиста.
                if _REPORT_END_LC and any(m in _norm_name(s) for m in _REPORT_END_LC):
                    report_ended = True
                continue

            # detail-строка ниже финальных итогов листа = служебная заметка финансиста → пропуск.
            if report_ended:
                continue

            if sum_f1 is None and sum_f2 is None and total is None:
                continue

            rows_out.append({
                "row": r_idx,
                "article": s,
                "group": current_group,
                "type_statii": config.determine_type(s, current_group),
                "sum_f1": sum_f1 or 0.0,
                "sum_f2": sum_f2 or 0.0,
                "total": total or 0.0,
                "comment": comment,
            })
        if rows_out:
            yield {
                "sheet_name": name,
                "rows": rows_out,
                "group_totals": group_totals,
                "general_totals": general_totals,
            }
