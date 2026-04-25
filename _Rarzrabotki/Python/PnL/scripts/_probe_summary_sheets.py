"""Проверить структуру сводных листов (PL_ЦО, PL_Логистика, Техника и др.)
- есть ли месячный заголовок в row 2, есть ли статьи в колонке B, есть ли суммы."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import config
from utils.excel_parser import find_month_columns
from openpyxl import load_workbook


def probe_sheet(ws, month_header):
    cols = find_month_columns(ws, month_header)
    if cols is None:
        return None, None, None
    ci_art, ci_f1, ci_f2, ci_tot, ci_pct, ci_cmt = cols
    # Scan B column for how many articles
    articles = 0
    first_article = None
    for r in range(config.PL_DATA_START_ROW, min(ws.max_row + 1, config.PL_DATA_START_ROW + 100)):
        v = ws.cell(r, ci_art + 1).value
        if v and str(v).strip():
            if articles == 0:
                first_article = str(v).strip()[:40]
            articles += 1
    return cols, articles, first_article


def main():
    for f in config.EXCEL_FILES:
        print(f"\n=== {f['label']} ===")
        wb = load_workbook(f["path"], data_only=True, read_only=False)
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            name = ws.title
            is_summary = name in config.SUMMARY_SHEET_EXACT or any(
                name.startswith(p) for p in config.SUMMARY_SHEET_PATTERNS
            )
            if not is_summary:
                continue  # интересуют только сводные
            cols, articles, first = probe_sheet(ws, f["month_header"])
            status = "NO-MONTH-COL" if cols is None else f"articles={articles} first='{first}'"
            print(f"  [SUMMARY] {name}: {status}")
        wb.close()


if __name__ == "__main__":
    main()
