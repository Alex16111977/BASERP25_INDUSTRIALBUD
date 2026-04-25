"""Re-inspect structure of Астарта. Тищенки after user's update."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook
import config

for f in config.EXCEL_FILES:
    print(f"\n=== {f['path']} ===")
    wb = load_workbook(f["path"], data_only=True)
    if "Астарта. Тищенки" not in wb.sheetnames:
        print("  Нет листа")
        continue
    ws = wb["Астарта. Тищенки"]
    # Print first 8 rows, columns A..CF
    for r in range(1, 8):
        for c in range(1, 90):
            v = ws.cell(r, c).value
            if v is None or v == "":
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            colletter = ws.cell(r, c).coordinate[:-len(str(r))]
            print(f"  row{r} {colletter}={str(v)[:40]!r}")
        print("---")
